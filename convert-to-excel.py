"""
Convertisseur de logs Fortigate vers format Excel créé par Hyperion.
"""


# Import Interface Graphique
import os.path
import re
import sys
import time
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory

# Bibliothèque de gestion des fichiers Excel
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def __load_file(filename: str) -> list:
    """
    Récupération du contenu du fichier de configuration
    :param filename : Chemin d'accès vers le fichier de configuration à convertir
    :return : Liste contenant toutes les lignes du fichier de configuration
    """
    file = open(filename, 'rb')
    brut = file.readlines()
    file.close()
    content = []
    for line in brut:
        content.append(line.decode('utf8'))
    return content


def check_file_already_exist(input_pathfile: str, output_path: str, prefix: str, overwrite: bool) -> str:
    """
    Vérification si le fichier de sortie existe déjà et ouvre une fenêtre de demande d'écrasement.
    Renomme le fichier sinon.
    :param input_pathfile : Chemin d'accès du fichier d'entrée
    :param output_path : Chemin d'accès du dossier de sortie
    :param prefix : Permet d'ajouter un préfixe au fichier pour distinguer les différents types de log
    :param overwrite : Doit-on écraser automatiquement les fichiers déjà existants
    :return : Chemin d'accès du fichier de sortie à écrire
    """
    input_name = os.path.splitext(os.path.basename(input_pathfile))[0]
    output_filename = "{}/output_{}_{}.xlsx".format(
        output_path,
        prefix,
        input_name
    )
    if os.path.isfile(output_filename):
        if not overwrite:
            answer = input("Voulez-vous écraser le fichier {} ? (y/n)".format(os.path.basename(output_filename)))
            if answer in ("y", "yes", "o", "oui"):
                print("Ecrasement du fichier {}".format(os.path.basename(output_filename)))
                return output_filename
        else:
            print("Ecrasement du fichier {}".format(os.path.basename(output_filename)))
            return output_filename
        j = 1
        output_filename = "{}/output_{} ({}).xlsx".format(
            output_path,
            input_name,
            j
        )

        while os.path.isfile(output_filename):
            print("Fichier {} déjà existant".format(output_filename))
            j += 1
            output_filename = "{}/output_{} ({}).xlsx".format(
                output_path,
                input_name,
                j
            )

    return output_filename


def __find_type(matches: list) -> str:
    for match in matches:
        if match[0] == "type":
            return match[1]
    return ""


def __find_sub_type(matches: list) -> str:
    for match in matches:
        if match[0] == "subtype":
            return match[1]
    return ""


def __init_sub_type(length: int) -> dict:
    data_init = {}
    for n in range(length - 2):
        data_init[str(n)] = []
    data_init['keys'] = ["{}".format(x) for x in range(length - 2)]
    return data_init


if __name__ == "__main__":
    # Récupération du fichier de config
    Tk().withdraw()
    input_file = askopenfilename()

    # Test du fichier avant ouverture
    if not os.path.isfile(input_file):
        print("Le fichier sélectionné est introuvable")
        sys.exit(2)

    # Emplacement de sortie
    output_dir = askdirectory(title="Emplacement de sortie")


    regex = r"\"([^=\"]+)=[\"]*([^=\"]+)[\"]*\""

    # Début de la conversion du fichier
    content = __load_file(input_file)
    data = {}

    print("Analyse du fichier en cours...")

    for line in content:
        items = line.split(",")
        data_tuples = []
        log_type = ""
        log_sub_type = ""

        for item in items:
            if item == "\"\"":
                data_tuples.append(("", ""))
            else:
                match = re.findall(regex, item)
                if len(match):
                    if match[0][0] == "type":
                        log_type = match[0][1]
                    elif match[0][0] == "subtype":
                        log_sub_type = match[0][1]
                    else:
                        data_tuples.append(match[0])

        if log_type not in data.keys():
            data[log_type] = {}
        if log_sub_type not in data[log_type].keys():
            data[log_type][log_sub_type] = __init_sub_type(len(items))

        for i in range(len(data_tuples)):
            if data_tuples[i][0]:
                if data_tuples[i][0] not in data[log_type][log_sub_type].keys():
                    data[log_type][log_sub_type][data_tuples[i][0]] = data[log_type][log_sub_type][str(i)]
                    del data[log_type][log_sub_type][str(i)]
                    data[log_type][log_sub_type]["keys"][i] = data_tuples[i][0]
                data[log_type][log_sub_type][data_tuples[i][0]].append(data_tuples[i][1])
            else:
                data[log_type][log_sub_type][data[log_type][log_sub_type]["keys"][i]].append("")

    print("Création du fichier Excel en cours...")
    # Ecriture des fichiers Excel
    for prefix in data.keys():
        output_file = check_file_already_exist(input_file, output_dir, prefix, False)

        # Création d'un nouveau fichier Excel
        wb = Workbook()
        wb.encoding = "cp1252"

        for sheet_name in data[prefix].keys():
            # Si le fichier existe (déjà écrit au moins une fois) récupération
            if os.path.isfile(output_file):
                wb = load_workbook(filename=output_file)

            print("Écriture de la feuille {}".format(sheet_name))
            ws = wb.create_sheet(sheet_name)

            # Récupération des colonnes valides
            cols = []
            for key in data[prefix][sheet_name]["keys"]:
                try:
                    int(key)
                except ValueError:
                    cols.append(key)

            # Ecriture des colonnes
            for k in range(len(cols)):
                c = ws.cell(row=1, column=k + 1, value=cols[k])
                c.font = Font(bold=True)
                c.border = thin_border

                # Ecriture des valeurs
                for j in range(len(data[prefix][sheet_name][cols[k]])):
                    ws.cell(row=j + 2, column=k + 1, value=data[prefix][sheet_name][cols[k]][j])

            # Freeze de la première ligne
            ws.freeze_panes = ws['B2']

            # Ecriture du fichier à chaque feuille pour économiser de la RAM
            print("Sauvegarde de la feuille {}".format(sheet_name))
            wb.save(output_file)

        wb.remove(wb["Sheet"])
        wb.save(output_file)
        print("Fichier {} créé avec succès. Cette fenêtre se fermera automatiquement dans 5 secondes.".format(output_file))
        time.sleep(5)
